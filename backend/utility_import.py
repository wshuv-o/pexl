"""Import an externally-generated utility Excel straight into the database.

The uploaded workbook already matches our utility-extraction schema (one row
per property × utility × billing period), e.g.:

    Property Name | Provider Name | Account | Service End Date |
    Service Start Date | Current Charges | Utility Type

This module parses that sheet, maps its columns onto the field names the rest
of the app (and the utility-template exporter) expects, creates a batch, and
bulk-inserts every row as a batch_record. No OCR, no PDF — the data is trusted
as-is.
"""

import io
import logging
import uuid

from fastapi import APIRouter, File, Form, UploadFile
from fastapi.responses import JSONResponse

import batch_service as bs

logger = logging.getLogger(__name__)
router = APIRouter()


# ---------------------------------------------------------------------------
# Column-header mapping (case/space/punctuation-insensitive)
# ---------------------------------------------------------------------------

def _norm(h: str) -> str:
    return "".join(ch for ch in str(h or "").lower() if ch.isalnum())


# normalized header -> canonical field name
_HEADER_MAP = {
    "propertyname": "property_name",
    "property": "property_name",
    "providername": "provider_name",
    "provider": "provider_name",
    "account": "account_number",
    "accountnumber": "account_number",
    "accountno": "account_number",
    "acct": "account_number",
    "serviceenddate": "service_end_date",
    "billingenddate": "service_end_date",
    "enddate": "service_end_date",
    "servicestartdate": "service_start_date",
    "billingstartdate": "service_start_date",
    "startdate": "service_start_date",
    "currentcharges": "amount",
    "amount": "amount",
    "charges": "amount",
    "totalamount": "amount",
    "utilitytype": "utility_type",
    "type": "utility_type",
    "utility": "utility_type",
    "address": "address",
}

# Utility type -> the total_* field the template groups on.
_UTILITY_FIELD = {
    "electricity": "total_electricity_bill",
    "electric": "total_electricity_bill",
    "gas": "total_gas_bill",
    "water": "total_water_bill",
    "sewer": "total_sewer_bill",
    "watersewer": "total_water_sewer_bill",
    "waterandsewer": "total_water_sewer_bill",
    "internet": "total_internet_bill",
    "phone": "total_phone_bill",
    "trash": "total_trash_bill",
}


def _to_str(v) -> str:
    if v is None:
        return ""
    # openpyxl gives datetimes for date cells — render as ISO date.
    if hasattr(v, "isoformat"):
        try:
            return v.date().isoformat() if hasattr(v, "date") else v.isoformat()
        except Exception:
            return str(v)
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _parse_rows(file_bytes: bytes):
    """Yield mapped field-dicts from the first worksheet.

    Returns (rows, headers, skipped_blank).
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = []
    headers = None
    col_field = {}  # column index -> canonical field
    skipped = 0
    for r_i, raw in enumerate(ws.iter_rows(values_only=True)):
        if headers is None:
            headers = [_to_str(c) for c in raw]
            for c_i, h in enumerate(raw):
                key = _HEADER_MAP.get(_norm(h))
                if key:
                    col_field[c_i] = key
            continue

        vals = {}
        for c_i, cell in enumerate(raw):
            field = col_field.get(c_i)
            if not field:
                continue
            s = _to_str(cell)
            if s:
                vals[field] = s

        # Drop fully blank rows (the file has trailing empties).
        if not any(vals.get(k) for k in ("property_name", "account_number", "amount", "utility_type")):
            skipped += 1
            continue

        # Map the amount onto the utility-type-specific total field so the
        # utility template groups it correctly. Keep the generic 'amount' too.
        utype = vals.get("utility_type", "")
        amount = vals.get("amount", "")
        if amount and utype:
            tf = _UTILITY_FIELD.get(_norm(utype))
            if tf:
                vals[tf] = amount
        # Use the service end date as the billing date the template buckets on.
        if vals.get("service_end_date"):
            vals.setdefault("billing_date", vals["service_end_date"])

        rows.append(vals)

    wb.close()
    return rows, headers, skipped


@router.post("/api/utility-import")
async def utility_import(
    file: UploadFile = File(...),
    name: str = Form(...),
    created_by: str = Form("import"),
):
    """Parse an uploaded utility Excel and insert every row into a new batch."""
    filename = file.filename or "utilities.xlsx"
    if not filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        return JSONResponse(status_code=400, content={"error": "Please upload an .xlsx file."})

    data = await file.read()
    if not data:
        return JSONResponse(status_code=400, content={"error": "Uploaded file is empty."})

    try:
        rows, headers, skipped = _parse_rows(data)
    except Exception as exc:
        logger.exception("utility-import parse failed")
        return JSONResponse(status_code=400, content={"error": f"Could not read Excel: {exc}"})

    if not rows:
        return JSONResponse(status_code=400, content={
            "error": "No data rows found. Detected headers: " + ", ".join(h for h in (headers or []) if h),
        })

    # One session_id for the whole import; page = 1-based row index (unique).
    session_id = "xls-" + uuid.uuid4().hex
    records = []
    for i, vals in enumerate(rows, start=1):
        # filename per row = property name so the utility template groups by
        # property; fall back to the workbook name.
        prop = vals.get("property_name") or filename
        records.append({
            "session_id": session_id,
            "filename": prop[:500],
            "page": i,
            "fields": vals,
        })

    try:
        batch_id = bs.create_batch(name, created_by, "utility_bill")
        inserted = bs.bulk_insert_records(batch_id, records)
    except Exception as exc:
        logger.exception("utility-import insert failed")
        return JSONResponse(status_code=500, content={"error": f"Database insert failed: {exc}"})

    return {
        "status": "ok",
        "batch_id": batch_id,
        "batch_name": name,
        "inserted": inserted,
        "skipped_blank": skipped,
        "detected_headers": [h for h in (headers or []) if h],
        "mapped_fields": sorted({k for r in rows for k in r.keys()}),
    }
