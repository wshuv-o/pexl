from __future__ import annotations

import concurrent.futures
import io
import logging
import os
import tempfile

import fitz
import pdfplumber
from fastapi import APIRouter
from fastapi.responses import JSONResponse, Response
from pydantic import BaseModel

from pdf_store import store, PDFDocument as _PDFDocument
from searchable_pdf import make_searchable_pdf

logger = logging.getLogger(__name__)

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    _EXCEL_AVAILABLE = True
except ImportError:
    _EXCEL_AVAILABLE = False

try:
    import camelot
    _CAMELOT_AVAILABLE = True
except ImportError:
    _CAMELOT_AVAILABLE = False

try:
    from docling.document_converter import DocumentConverter as _DoclingConverter, PdfFormatOption as _DocPdfFmtOpt
    from docling.datamodel.pipeline_options import PdfPipelineOptions as _DocPdfOpts, TableFormerMode as _TableFormerMode
    from docling.datamodel.base_models import InputFormat as _DocInputFormat
    _DOCLING_AVAILABLE = True
except ImportError:
    _DOCLING_AVAILABLE = False

# Per-page Camelot cache: key = "session_id:page", value = (lattice_tables, stream_tables).
# Camelot table objects are fully in-memory after read_pdf() returns; caching them means
# Ghostscript runs only once per page instead of on every region click.
import threading as _threading
_camelot_page_cache: dict[str, tuple[list, list]] = {}
_camelot_cache_lock = _threading.Lock()
_CAMELOT_CACHE_MAX = 60  # max (session, page) pairs to keep


# Compat shim: the old startup-time `set_ocr_engine(engine)` hook is no
# longer needed — table_extract now pulls the engine lazily via
# ocr_lazy.get_ocr_engine() at the point of use. Kept as a no-op so any
# external caller that still invokes it doesn't crash.
def set_ocr_engine(_engine) -> None:  # pragma: no cover
    pass


def _get_ocr_engine():
    """Lazy accessor — imports here to avoid a circular import with main."""
    from ocr_lazy import get_ocr_engine
    return get_ocr_engine()


router = APIRouter()

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _bbox_overlap_ratio(b1, b2) -> float:
    if b1 is None or b2 is None:
        return 0.0
    ix1, iy1 = max(b1[0], b2[0]), max(b1[1], b2[1])
    ix2, iy2 = min(b1[2], b2[2]), min(b1[3], b2[3])
    if ix2 <= ix1 or iy2 <= iy1:
        return 0.0
    inter = (ix2 - ix1) * (iy2 - iy1)
    a1 = max((b1[2] - b1[0]) * (b1[3] - b1[1]), 1)
    a2 = max((b2[2] - b2[0]) * (b2[3] - b2[1]), 1)
    return inter / min(a1, a2)


def _split_multiline_df(df: "pd.DataFrame") -> "pd.DataFrame":
    new_rows: list[list[str]] = []
    for _, row in df.iterrows():
        cells = [str(c) if c is not None else "" for c in row]
        split = [c.split("\n") for c in cells]
        max_lines = max(len(s) for s in split)
        if max_lines <= 1:
            new_rows.append(cells)
        else:
            padded = [s + [""] * (max_lines - len(s)) for s in split]
            for line_idx in range(max_lines):
                line = [padded[col][line_idx].strip() for col in range(len(cells))]
                if any(line):
                    new_rows.append(line)
    if not new_rows:
        return df
    max_cols = max(len(r) for r in new_rows)
    padded_rows = [r + [""] * (max_cols - len(r)) for r in new_rows]
    return pd.DataFrame(padded_rows)


def _words_to_df(words: list) -> "pd.DataFrame | None":
    """Convert word list [(x0,y0,x1,y1,text), ...] to a structured DataFrame.

    Strategy:
      • Row clustering by mid-y using an adaptive y-tolerance derived from
        the median word height.
      • Column detection via **vertical whitespace projection**: build a
        coverage histogram across x, find vertical gaps where no word
        overlaps, and place column boundaries in the middle of each gap.
        This handles left-, right-, and center-aligned columns equally
        well, unlike gap-median heuristics.
    """
    if not words:
        return None

    items: list[tuple[float, float, float, float, str]] = []
    for w in words:
        if len(w) < 5:
            continue
        text = str(w[4]).strip()
        if text:
            items.append((float(w[0]), float(w[1]), float(w[2]), float(w[3]), text))

    if len(items) < 2:
        return None

    # ── Step 1: row clustering by mid-y ──────────────────────────────────────
    heights = sorted(it[3] - it[1] for it in items if it[3] > it[1])
    if not heights:
        return None
    # GAP-based row clustering: new row when the mid-y gap between consecutive
    # words exceeds half a median text height. Splits visually-separate lines
    # cleanly even with OCR jitter on individual word baselines — satisfies
    # the "identify new line" requirement of the OCR → table-region pipeline.
    median_h = heights[len(heights) // 2]
    line_gap = max(median_h * 0.5, 2.0)

    with_mid = sorted(((it, (it[1] + it[3]) / 2.0) for it in items), key=lambda x: x[1])
    rows: list[list[tuple[float, float, float, float, str]]] = [[with_mid[0][0]]]
    prev_mid = with_mid[0][1]
    for it, mid_y in with_mid[1:]:
        if mid_y - prev_mid <= line_gap:
            rows[-1].append(it)
        else:
            rows.append([it])
        prev_mid = mid_y

    rows = [sorted(r, key=lambda w: w[0]) for r in rows]

    # ── Step 2: column boundaries via whitespace projection ──────────────────
    min_x = min(it[0] for it in items)
    max_x = max(it[2] for it in items)
    width = max_x - min_x
    if width <= 0:
        return pd.DataFrame([[it[4] for it in items]])

    # Header rows like "Period From 1/1/2023 To 12/31/2023" can span the full
    # table width — including them in the histogram fills every bin and
    # destroys the white-space channels between data columns. Use only items
    # narrower than 45% of the table width to build the column profile, but
    # still place every word (including the wide ones) into the result.
    detect_items = [it for it in items if (it[2] - it[0]) < width * 0.45]
    if len(detect_items) < max(4, len(items) // 4):
        detect_items = items

    bin_size = 0.5  # 0.5pt bins
    n_bins = int(width / bin_size) + 2
    profile = [0] * n_bins
    for it in detect_items:
        i0 = max(0, int((it[0] - min_x) / bin_size))
        i1 = min(n_bins - 1, int((it[2] - min_x) / bin_size))
        for i in range(i0, i1 + 1):
            profile[i] += 1

    # Find runs of zeros (white-space "channels") — these are column gutters.
    # Require gap ≥ 40% of median text height to avoid splitting inside words.
    min_gap_pts = max(median_h * 0.4, 2.0)
    min_gap_bins = max(int(min_gap_pts / bin_size), 2)

    boundaries: list[float] = []
    i = 0
    while i < n_bins:
        if profile[i] == 0:
            start = i
            while i < n_bins and profile[i] == 0:
                i += 1
            # Only count gaps interior to the range (drop leading/trailing)
            if start > 0 and i < n_bins and (i - start) >= min_gap_bins:
                boundaries.append(min_x + (start + i) / 2.0 * bin_size)
        else:
            i += 1

    # Right-edge clustering on numeric tokens — always run, not just as fallback.
    # Financial tables right-align AMOUNT/BALANCE numbers so their x1 coordinates
    # cluster tightly at column right-edges. The whitespace projection often misses
    # the boundary between a wide text column (SPLIT) and the adjacent numeric
    # columns (AMOUNT, BALANCE) because long SPLIT text reduces the gap below the
    # threshold. We compute numeric-column boundaries independently and merge any
    # that the projection missed.
    numeric_items = [it for it in items if _is_numeric_token(it[4])]
    if len(numeric_items) >= 3:
        right_edges = sorted(it[2] for it in numeric_items)
        re_clusters: list[list[float]] = [[right_edges[0]]]
        re_tol = max(median_h * 1.2, 6.0)
        for x in right_edges[1:]:
            if x - re_clusters[-1][-1] <= re_tol:
                re_clusters[-1].append(x)
            else:
                re_clusters.append([x])
        min_support = max(2, len(numeric_items) // 30)
        col_right_edges = sorted(
            sum(cl) / len(cl) for cl in re_clusters if len(cl) >= min_support
        )
        if len(col_right_edges) >= 2:
            # Build midpoints between consecutive right-edge clusters as candidate
            # column boundaries, then add any that are not already covered by the
            # projection result (within a 1.5× text-height merge tolerance).
            _merge_tol = max(median_h * 1.5, 8.0)
            for j in range(len(col_right_edges) - 1):
                _nb = (col_right_edges[j] + col_right_edges[j + 1]) / 2.0
                if not any(abs(_nb - pb) <= _merge_tol for pb in boundaries):
                    boundaries.append(_nb)
            boundaries.sort()

    n_cols = len(boundaries) + 1

    def _col(x0: float, x1: float, text: str) -> int:
        # Numeric tokens are right-aligned — anchor on their right edge so they
        # land in the correct numeric column.  Text tokens are left-aligned —
        # anchor on their left edge so long words don't bleed into the next col.
        anchor = x1 if _is_numeric_token(text) else x0
        for idx, b in enumerate(boundaries):
            if anchor <= b:
                return idx
        return n_cols - 1

    # ── Step 3: assign words to (row, col) cells ─────────────────────────────
    rows_data: list[list[str]] = []
    for row in rows:
        cells = [""] * n_cols
        for _x0, _y0, x1, _y1, text in row:
            ci = _col(_x0, x1, text)
            cells[ci] = (cells[ci] + " " + text).strip() if cells[ci] else text
        if any(c for c in cells):
            rows_data.append(cells)

    if not rows_data:
        return None

    return pd.DataFrame(rows_data)


def _is_numeric_token(text: str) -> bool:
    """Heuristic: is this word a financial number like '423,820.73' or '-1,234'?"""
    s = (text or "").strip()
    if not s:
        return False
    if not any(c.isdigit() for c in s):
        return False
    cleaned = s.replace(",", "").replace(".", "").replace("-", "").replace("(", "").replace(")", "").replace("$", "").replace("%", "")
    return cleaned.isdigit()


def _apply_excel_formatting(
    excel_bytes: bytes,
    sheet_regions: dict[str, list[tuple[int, int, int]]],
) -> bytes:
    wb = load_workbook(io.BytesIO(excel_bytes))

    _thin       = Side(style="thin")
    _border     = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _hdr_fill   = PatternFill("solid", fgColor="D9E1F2")
    _hdr_font   = Font(bold=True)
    _title_font = Font(bold=True, size=12)
    _wrap       = Alignment(wrap_text=True, vertical="top")

    for sheet_name, regions in sheet_regions.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        if ws["A1"].value:
            ws["A1"].font = _title_font

        for (start_row, n_rows, n_cols) in regions:
            for r in range(start_row, start_row + n_rows):
                for c in range(1, n_cols + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.border = _border
                    cell.alignment = _wrap
                    if r == start_row:
                        cell.font = _hdr_font
                        cell.fill = _hdr_fill

        for col_cells in ws.columns:
            max_len = 0
            for cell in col_cells:
                try:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value).split("\n")[0]))
                except Exception:
                    pass
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Region extraction helpers
# ---------------------------------------------------------------------------

class RegionRequest(BaseModel):
    page: int
    x: float
    y: float
    width: float
    height: float
    rows: list[list[str]] | None = None  # pre-extracted rows; skip re-extraction when present
    transposed: bool = False


def _get_pdf_bytes_for_session(session_id: str, target_page: int = 0):
    """Return (pdf_doc, bytes_with_text_layer) for a session.

    If the full searchable PDF is already cached, return it immediately.
    If OCR is needed and target_page is given, OCR only that one page so the
    caller doesn't wait for the whole document to be processed — typically
    reduces wait time from minutes to a few seconds.
    """
    pdf_doc = store.get(session_id)
    if not pdf_doc:
        return None, None

    # Full searchable PDF already available (background task completed).
    if pdf_doc.searchable_pdf_bytes:
        logger.info("Region: using cached full searchable PDF")
        return pdf_doc, pdf_doc.searchable_pdf_bytes

    needs_ocr = any(p.get("is_ocr") or p.get("is_vector") for p in pdf_doc.page_info)
    if not needs_ocr or _get_ocr_engine() is None:
        logger.info("Region: using native PDF bytes (no OCR needed)")
        return pdf_doc, pdf_doc.pdf_bytes

    # OCR needed but full doc not ready yet. OCR just the requested page.
    if target_page > 0 and target_page <= len(pdf_doc.page_info):
        logger.info("Region: single-page OCR for page %d (full doc not ready)", target_page)
        try:
            raw = pdf_doc.pdf_bytes
            if raw is None:
                return pdf_doc, None
            _single = fitz.open()
            with fitz.open(stream=raw, filetype="pdf") as _full:
                _single.insert_pdf(_full, from_page=target_page - 1, to_page=target_page - 1)
            _single_bytes = _single.tobytes()
            _single.close()

            _temp_doc = _PDFDocument(_single_bytes, f"page_{target_page}")
            _temp_doc.page_info = [pdf_doc.page_info[target_page - 1]]
            # Use cached OCR words for this page if background task already did it.
            if pdf_doc.structured_pages and len(pdf_doc.structured_pages) >= target_page:
                _temp_doc.structured_pages = [pdf_doc.structured_pages[target_page - 1]]

            _single_searchable = make_searchable_pdf(_temp_doc, _get_ocr_engine(), image_only=True)
            _temp_doc.close()

            if _single_searchable:
                logger.info("Region: single-page OCR complete (%d bytes)", len(_single_searchable))
                return pdf_doc, _single_searchable
        except Exception as exc:
            logger.warning("Region: single-page OCR failed: %s", exc)

    # Fallback: use native bytes (extraction may find nothing on scanned pages).
    logger.info("Region: falling back to native PDF bytes")
    return pdf_doc, pdf_doc.pdf_bytes


# ---------------------------------------------------------------------------
# Geometric cell reconstruction
#
# Many PDFs contain *actual* ruling lines for table borders. When they do, the
# correct way to recover the table is to find those lines, build a grid from
# them, and walk cell-by-cell — exactly preserving merges, spans, and the
# original row/column structure. Whitespace clustering only kicks in when no
# such geometry exists.
# ---------------------------------------------------------------------------

def _cluster_coords(coords: list[float], tol: float = 3.0) -> list[float]:
    """Cluster nearby 1-D coordinates into their averaged centers."""
    if not coords:
        return []
    coords = sorted(coords)
    clusters: list[list[float]] = [[coords[0]]]
    for c in coords[1:]:
        if c - clusters[-1][-1] <= tol:
            clusters[-1].append(c)
        else:
            clusters.append([c])
    return [sum(cl) / len(cl) for cl in clusters]


def _collect_segments_in_clip(page, clip):
    """Walk page.get_drawings() and return ruling lines that fall inside clip.

    Returns (h_segments, v_segments):
      h_segments: list of (x_left, x_right, y)
      v_segments: list of (y_top, y_bot, x)
    Both horizontal/vertical thin rectangles and explicit line items count.
    """
    h_segments: list[tuple[float, float, float]] = []
    v_segments: list[tuple[float, float, float]] = []
    try:
        drawings = page.get_drawings()
    except Exception:
        return h_segments, v_segments

    cx0, cy0, cx1, cy1 = clip.x0, clip.y0, clip.x1, clip.y1

    def _add_h(xa: float, xb: float, y: float) -> None:
        if y < cy0 - 1 or y > cy1 + 1:
            return
        if xb < cx0 - 1 or xa > cx1 + 1:
            return
        h_segments.append((max(xa, cx0), min(xb, cx1), y))

    def _add_v(ya: float, yb: float, x: float) -> None:
        if x < cx0 - 1 or x > cx1 + 1:
            return
        if yb < cy0 - 1 or ya > cy1 + 1:
            return
        v_segments.append((max(ya, cy0), min(yb, cy1), x))

    for d in drawings:
        for item in d.get("items", []):
            op = item[0]
            if op == "l" and len(item) >= 3:
                p1, p2 = item[1], item[2]
                x1, y1 = float(p1.x), float(p1.y)
                x2, y2 = float(p2.x), float(p2.y)
                if abs(y1 - y2) <= 0.8 and abs(x1 - x2) >= 1.0:
                    _add_h(min(x1, x2), max(x1, x2), (y1 + y2) / 2.0)
                elif abs(x1 - x2) <= 0.8 and abs(y1 - y2) >= 1.0:
                    _add_v(min(y1, y2), max(y1, y2), (x1 + x2) / 2.0)
            elif op == "re" and len(item) >= 2:
                r = item[1]
                x0, y0 = float(r.x0), float(r.y0)
                x1, y1 = float(r.x1), float(r.y1)
                w, h = x1 - x0, y1 - y0
                # Very thin rectangles are border strokes drawn as fills
                if w >= 1.0 and h <= 1.5:
                    _add_h(x0, x1, (y0 + y1) / 2.0)
                elif h >= 1.0 and w <= 1.5:
                    _add_v(y0, y1, (x0 + x1) / 2.0)
                else:
                    # Normal rectangle: contributes its 4 edges
                    _add_h(x0, x1, y0)
                    _add_h(x0, x1, y1)
                    _add_v(y0, y1, x0)
                    _add_v(y0, y1, x1)
    return h_segments, v_segments


def _line_coverage(segments, axis_val: float, perp_lo: float, perp_hi: float, axis_tol: float = 2.0) -> float:
    """Sum overlap of all segments near axis_val with the interval [perp_lo, perp_hi]."""
    cov = 0.0
    for s_lo, s_hi, s_axis in segments:
        if abs(s_axis - axis_val) <= axis_tol:
            o = min(s_hi, perp_hi) - max(s_lo, perp_lo)
            if o > 0:
                cov += o
    return cov


def _extract_cells_from_geometry(page, clip):
    """Reconstruct merged cells from actual ruling lines.

    Returns {"cells": [...], "n_rows": int, "n_cols": int} or None.

    Each cell dict:
        {row, col, rowspan, colspan, text, bbox: [x0,y0,x1,y1]}
    """
    h_segs, v_segs = _collect_segments_in_clip(page, clip)
    if len(h_segs) < 2 or len(v_segs) < 2:
        return None

    h_ys = _cluster_coords([s[2] for s in h_segs], tol=3.0)
    v_xs = _cluster_coords([s[2] for s in v_segs], tol=3.0)
    if len(h_ys) < 2 or len(v_xs) < 2:
        return None

    table_x0, table_x1 = min(v_xs), max(v_xs)
    table_y0, table_y1 = min(h_ys), max(h_ys)
    table_w = table_x1 - table_x0
    table_h = table_y1 - table_y0
    if table_w < 20.0 or table_h < 20.0:
        return None

    # Keep only dividers whose lines actually span a meaningful portion of the table.
    row_dividers = [y for y in h_ys if _line_coverage(h_segs, y, table_x0, table_x1) >= 0.4 * table_w]
    col_dividers = [x for x in v_xs if _line_coverage(v_segs, x, table_y0, table_y1) >= 0.4 * table_h]

    if len(row_dividers) < 2 or len(col_dividers) < 2:
        return None

    row_dividers.sort()
    col_dividers.sort()
    n_rows = len(row_dividers) - 1
    n_cols = len(col_dividers) - 1
    if n_rows < 1 or n_cols < 1:
        return None

    # Union-find: merge grid cells across edges that don't have a real line.
    parent = list(range(n_rows * n_cols))

    def find(i: int) -> int:
        while parent[i] != i:
            parent[i] = parent[parent[i]]
            i = parent[i]
        return i

    def union(a: int, b: int) -> None:
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb

    def has_h(y: float, x_left: float, x_right: float) -> bool:
        span = x_right - x_left
        return span <= 0 or _line_coverage(h_segs, y, x_left, x_right) >= 0.6 * span

    def has_v(x: float, y_top: float, y_bot: float) -> bool:
        span = y_bot - y_top
        return span <= 0 or _line_coverage(v_segs, x, y_top, y_bot) >= 0.6 * span

    for r in range(n_rows):
        for c in range(n_cols):
            idx = r * n_cols + c
            if c + 1 < n_cols and not has_v(col_dividers[c + 1], row_dividers[r], row_dividers[r + 1]):
                union(idx, idx + 1)
            if r + 1 < n_rows and not has_h(row_dividers[r + 1], col_dividers[c], col_dividers[c + 1]):
                union(idx, idx + n_cols)

    groups: dict[int, list[tuple[int, int]]] = {}
    for r in range(n_rows):
        for c in range(n_cols):
            groups.setdefault(find(r * n_cols + c), []).append((r, c))

    cells_out: list[dict] = []
    emitted: set[int] = set()
    for r in range(n_rows):
        for c in range(n_cols):
            root = find(r * n_cols + c)
            if root in emitted:
                continue
            emitted.add(root)
            members = groups[root]
            rs = [m[0] for m in members]
            cs = [m[1] for m in members]
            r_min, r_max = min(rs), max(rs)
            c_min, c_max = min(cs), max(cs)
            x0 = col_dividers[c_min]
            y0 = row_dividers[r_min]
            x1 = col_dividers[c_max + 1]
            y1 = row_dividers[r_max + 1]

            try:
                words = page.get_text("words", clip=fitz.Rect(x0, y0, x1, y1))
            except Exception:
                words = []

            text = _words_to_cell_text(words)

            cells_out.append({
                "row": r_min,
                "col": c_min,
                "rowspan": r_max - r_min + 1,
                "colspan": c_max - c_min + 1,
                "text": text,
                "bbox": [x0, y0, x1, y1],
            })

    return {"cells": cells_out, "n_rows": n_rows, "n_cols": n_cols}


def _words_to_cell_text(words: list) -> str:
    """Join words inside a cell into multi-line text, preserving reading order."""
    if not words:
        return ""
    items = [(float(w[0]), float(w[1]), float(w[3]), str(w[4]).strip()) for w in words if len(w) >= 5 and str(w[4]).strip()]
    if not items:
        return ""
    heights = sorted(t[2] - t[1] for t in items if t[2] > t[1])
    median_h = heights[len(heights) // 2] if heights else 10.0
    y_tol = max(median_h * 0.6, 2.5)

    items.sort(key=lambda t: (t[1], t[0]))
    lines: list[list[tuple[float, str]]] = [[(items[0][0], items[0][3])]]
    anchor = items[0][1]
    for x0, y0, _y1, txt in items[1:]:
        if abs(y0 - anchor) <= y_tol:
            lines[-1].append((x0, txt))
        else:
            lines.append([(x0, txt)])
            anchor = y0
    return "\n".join(" ".join(t for _x, t in sorted(line, key=lambda p: p[0])) for line in lines)


def _cells_to_flat_df(cells: list[dict], n_rows: int, n_cols: int) -> "pd.DataFrame":
    """Flatten a cell list into a 2-D DataFrame. Spanned positions are blank."""
    grid = [["" for _ in range(n_cols)] for _ in range(n_rows)]
    for cell in cells:
        grid[cell["row"]][cell["col"]] = cell["text"]
    return pd.DataFrame(grid)


def _words_to_df_with_boundaries(
    pdf_bytes: bytes,
    page: int,
    rx1: float,
    ry1_top: float,
    rx2: float,
    ry2_bot: float,
    col_boundaries: list[float],
) -> "pd.DataFrame | None":
    """Bucket every word inside the user's selection into pre-computed columns.

    ``col_boundaries`` are interior x-edges (PDF points, top-left origin)
    that came from a full-page camelot detection — these are far more
    reliable than column detection on the limited word set inside the
    selection. Rows are clustered by y-bbox overlap so OCR'd text with
    vertical jitter still groups properly. Result respects the selection
    extent exactly: every word inside the box ends up in the output.
    """
    try:
        with fitz.open(stream=pdf_bytes, filetype="pdf") as fitz_doc:
            pg = fitz_doc[page - 1]
            clip = fitz.Rect(rx1, ry1_top, rx2, ry2_bot)
            raw = pg.get_text("words", clip=clip)
    except Exception:
        return None
    if not raw:
        return None

    items = [
        (float(w[0]), float(w[1]), float(w[2]), float(w[3]), str(w[4]).strip())
        for w in raw if str(w[4]).strip()
    ]
    if not items:
        return None

    # Row clustering: GAP-based on mid-y. A new row starts whenever the
    # mid-y gap between consecutive (sorted) words exceeds half a median
    # text height. This naturally splits visually-separate lines even
    # when OCR jitter would have confused an overlap-based test. It also
    # respects "identify new line" semantics — every distinct OCR line in
    # the selection becomes its own row.
    heights = sorted(it[3] - it[1] for it in items if it[3] > it[1])
    if not heights:
        return None
    median_h = heights[len(heights) // 2]
    line_gap = max(median_h * 0.5, 2.0)

    with_mid = sorted(((it, (it[1] + it[3]) / 2.0) for it in items), key=lambda x: x[1])
    rows: list[list[tuple[float, float, float, float, str]]] = [[with_mid[0][0]]]
    prev_mid = with_mid[0][1]
    for it, mid_y in with_mid[1:]:
        if mid_y - prev_mid <= line_gap:
            rows[-1].append(it)
        else:
            rows.append([it])
        prev_mid = mid_y

    n_cols = len(col_boundaries) + 1

    def _col_for(x0: float, x1: float) -> int:
        center = (x0 + x1) / 2.0
        for i, b in enumerate(col_boundaries):
            if center < b:
                return i
        return n_cols - 1

    rows_data: list[list[str]] = []
    for r in rows:
        cells = [""] * n_cols
        for x0, _y0, x1, _y1, text in sorted(r, key=lambda t: t[0]):
            ci = _col_for(x0, x1)
            cells[ci] = (cells[ci] + " " + text).strip() if cells[ci] else text
        if any(c for c in cells):
            rows_data.append(cells)

    if not rows_data:
        return None
    return pd.DataFrame(rows_data)


def _score_df(df: "pd.DataFrame | None") -> float:
    """Score a candidate DataFrame: prefer balanced shapes over 1xN / Nx1.

    A result that is all rows in one column (n_cols=1) or all columns in
    one row (n_rows=1) gets penalized to roughly 1/10 of its cell count.
    Otherwise the score is rows × cols, with a small bonus for >=2 cols.
    """
    if df is None or df.empty:
        return -1.0
    n_rows, n_cols = df.shape
    if n_rows == 0 or n_cols == 0:
        return -1.0
    # Count non-empty cells (real signal, not padding)
    non_empty = sum(1 for row in df.values.tolist() for c in row if str(c).strip())
    if non_empty == 0:
        return -1.0
    if n_cols < 2 or n_rows < 2:
        return non_empty * 0.1
    # Reward balanced tables; heavy penalty for very skewed shapes
    aspect = min(n_rows, n_cols) / max(n_rows, n_cols)
    return non_empty * (1.0 + aspect)


def _region_ocr_extract(
    pdf_bytes: bytes,
    eff_page: int,
    rx1: float,
    ry1_top: float,
    rx2: float,
    ry2_bot: float,
) -> "pd.DataFrame | None":
    """Render the selected region and run RapidOCR on it directly.

    Used when force_ocr=True: the native text layer is ignored entirely —
    fitz renders the original scan pixels, OCR reads them, and _words_to_df
    builds the column structure. Works immediately without waiting for the
    full-document background OCR to finish.
    """
    _engine = _get_ocr_engine()
    if _engine is None or not _EXCEL_AVAILABLE:
        return None

    _DPI = 200  # higher than full-page 150 dpi — small regions need more detail
    _scale = _DPI / 72.0

    try:
        import numpy as _np
        with fitz.open(stream=pdf_bytes, filetype="pdf") as _fdoc:
            _pg = _fdoc[eff_page - 1]
            _clip = fitz.Rect(rx1, ry1_top, rx2, ry2_bot)
            _pix = _pg.get_pixmap(matrix=fitz.Matrix(_scale, _scale), clip=_clip, colorspace=fitz.csRGB)

        if _pix.width < 20 or _pix.height < 20:
            return None

        _img_arr = _np.frombuffer(_pix.samples, dtype=_np.uint8).reshape(_pix.height, _pix.width, 3).copy()
        _pix = None

        try:
            _results = _engine.ocr(_img_arr, cls=True)
        except Exception as exc:
            logger.debug("Region OCR engine failed: %s", exc)
            return None
        finally:
            _img_arr = None

        if not _results or not _results[0]:
            return None

        words: list[tuple[float, float, float, float, str]] = []
        for _item in _results[0]:
            try:
                _poly, (_text, _conf) = _item[0], _item[1]
                _text = (_text or "").strip()
                if not _text:
                    continue
                _xs = [p[0] for p in _poly]
                _ys = [p[1] for p in _poly]
                # Convert pixel coords → PDF points and shift by region origin
                words.append((
                    min(_xs) / _scale + rx1,
                    min(_ys) / _scale + ry1_top,
                    max(_xs) / _scale + rx1,
                    max(_ys) / _scale + ry1_top,
                    _text,
                ))
            except Exception:
                continue

        if not words:
            return None

        logger.info("Region OCR: %d words from image", len(words))
        return _words_to_df(words)

    except Exception as exc:
        logger.debug("Region OCR extract failed: %s", exc)
        return None


def _docling_table_extract(
    pdf_bytes: bytes,
    eff_page: int,
    rx1: float,
    ry1_top: float,
    rx2: float,
    ry2_bot: float,
) -> "list[pd.DataFrame]":
    """Embed the clipped region into a single-page image-PDF and run Docling
    TableFormer (ACCURATE mode) on it via the full PDF pipeline.

    Using the PDF pipeline (rather than the image pipeline) gives access to
    TableFormer's structure recognition, which handles both bordered and
    borderless tables. Only called on scanned/non-vector pages.
    Returns a list of DataFrames (one per detected table), possibly empty.
    """
    if not _DOCLING_AVAILABLE or not _EXCEL_AVAILABLE:
        return []

    _DPI = 300  # high-res crop — small regions need detail for the model
    _scale = _DPI / 72.0
    tmp_pdf_path: str | None = None
    try:
        with fitz.open(stream=pdf_bytes, filetype="pdf") as _fdoc:
            _pg = _fdoc[eff_page - 1]
            _clip = fitz.Rect(rx1, ry1_top, rx2, ry2_bot)
            _pix = _pg.get_pixmap(matrix=fitz.Matrix(_scale, _scale), clip=_clip, colorspace=fitz.csRGB)

        if _pix.width < 20 or _pix.height < 20:
            return []

        # Embed the rendered pixels into a fresh single-page PDF so we can
        # use Docling's PDF pipeline (which has TableFormer) instead of the
        # image pipeline (which lacks it).
        _img_w_pts = _pix.width * 72.0 / _DPI
        _img_h_pts = _pix.height * 72.0 / _DPI
        _new_doc = fitz.open()
        _new_pg = _new_doc.new_page(width=_img_w_pts, height=_img_h_pts)
        _new_pg.insert_image(fitz.Rect(0, 0, _img_w_pts, _img_h_pts), stream=_pix.tobytes("png"))
        _img_pdf_bytes = _new_doc.tobytes()
        _new_doc.close()
        _pix = None

        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as _fh:
            _fh.write(_img_pdf_bytes)
            tmp_pdf_path = _fh.name

        # PDF pipeline with TableFormer ACCURATE
        _pipe_opts = _DocPdfOpts(do_table_structure=True)
        _pipe_opts.table_structure_options.mode = _TableFormerMode.ACCURATE
        _converter = _DoclingConverter(
            format_options={_DocInputFormat.PDF: _DocPdfFmtOpt(pipeline_options=_pipe_opts)}
        )
        _result = _converter.convert(tmp_pdf_path)

        dfs: list[pd.DataFrame] = []
        for _tbl in _result.document.tables:
            try:
                _df = _tbl.export_to_dataframe()
                if _df is None or _df.empty:
                    continue
                if isinstance(_df.columns, pd.MultiIndex):
                    _df.columns = [" ".join(str(c) for c in col).strip() for col in _df.columns]
                _df = _df.reset_index(drop=True)
                dfs.append(_df)
            except Exception:
                pass
        return dfs

    except Exception as exc:
        logger.debug("Docling table extract failed: %s", exc)
        return []

    finally:
        if tmp_pdf_path:
            try:
                os.unlink(tmp_pdf_path)
            except Exception:
                pass


def _extract_region_table(pdf_bytes: bytes, page: int, x: float, y: float, width: float, height: float, session_id: str = ""):
    """Extract a table from a normalized region.

    Pipeline:
      0. Geometric reconstruction from page.get_drawings() — if real ruling
         lines exist, build the grid directly. Preserves merged cells exactly.
      1. camelot lattice  — bordered/grid tables (no merge info)
      2. camelot stream   — borderless structured tables
      3. pdfplumber lines — line-geometry fallback
      4. fitz word-bucket — whitespace-projection clustering

    Returns a dict {df, cells, n_rows, n_cols, source} or None.
    `cells` is a list of cell dicts with rowspan/colspan when geometry
    succeeded; None for the other strategies.
    """
    with fitz.open(stream=pdf_bytes, filetype="pdf") as _dim_doc:
        n_doc_pages = len(_dim_doc)
        # When pdf_bytes is a single-page extract for the target page (from
        # single-page OCR in _get_pdf_bytes_for_session), map page → 1.
        eff_page = 1 if n_doc_pages == 1 else page
        if eff_page < 1 or eff_page > n_doc_pages:
            return None
        _pg = _dim_doc[eff_page - 1]
        W, H = _pg.rect.width, _pg.rect.height

    rx1 = x * W
    ry1_top = y * H
    rx2 = (x + width) * W
    ry2_bot = (y + height) * H

    # ── Pass 0: geometric reconstruction from ruling lines ───────────────────
    try:
        with fitz.open(stream=pdf_bytes, filetype="pdf") as fitz_doc:
            pg = fitz_doc[eff_page - 1]
            clip = fitz.Rect(rx1, ry1_top, rx2, ry2_bot)
            geom = _extract_cells_from_geometry(pg, clip)
    except Exception as exc:
        logger.debug("geometric reconstruction failed: %s", exc)
        geom = None

    if geom and geom["n_rows"] >= 1 and geom["n_cols"] >= 1:
        df = _cells_to_flat_df(geom["cells"], geom["n_rows"], geom["n_cols"])
        score = _score_df(df)
        # Require a real >=2x2 grid AND that real content lands in >=2 columns.
        # Otherwise the geometric "grid" is just rows of full-width text that
        # all collapse into column 0 — let the smarter word fallback handle it.
        non_empty_per_col = [0] * geom["n_cols"]
        for cell in geom["cells"]:
            if str(cell.get("text", "")).strip():
                non_empty_per_col[cell["col"]] += 1
        filled_cols = sum(1 for n in non_empty_per_col if n > 0)

        if score > 0 and geom["n_rows"] >= 2 and geom["n_cols"] >= 2 and filled_cols >= 2:
            logger.info(
                "Region: geometry → %dx%d with %d merged cells, %d filled cols",
                geom["n_rows"], geom["n_cols"], len(geom["cells"]), filled_cols,
            )
            return {
                "df": df,
                "cells": geom["cells"],
                "n_rows": geom["n_rows"],
                "n_cols": geom["n_cols"],
                "source": "geometry",
            }

    # camelot uses bottom-left origin
    camelot_area = f"{rx1:.2f},{H - ry1_top:.2f},{rx2:.2f},{H - ry2_bot:.2f}"
    # Same region in camelot's bottom-left coordinate space, used to test
    # overlap with table bboxes detected by the full-page passes below.
    sel_bbox_bl = (rx1, H - ry2_bot, rx2, H - ry1_top)

    candidates: list[tuple[str, "pd.DataFrame"]] = []

    # Extract only the target page into a 1-page PDF so Ghostscript renders
    # exactly one page instead of the entire document. This is the main
    # perf fix: camelot passes pages="1" on the single-page file.
    _single_doc = fitz.open()
    with fitz.open(stream=pdf_bytes, filetype="pdf") as _full:
        _single_doc.insert_pdf(_full, from_page=eff_page - 1, to_page=eff_page - 1)
    _single_page_bytes = _single_doc.tobytes()
    _single_doc.close()

    tmp_path: str | None = None
    _has_drawings: bool = False
    try:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as fh:
            fh.write(_single_page_bytes)
            tmp_path = fh.name

        # Only run Camelot when the page has vector drawing lines — those are
        # the table borders Camelot needs to detect a grid. Scanned pages and
        # pure-text pages have no drawings; Camelot would burn Ghostscript time
        # and return nothing. OCR overlays add text but not drawings, so a
        # searchable scanned PDF is correctly skipped here too.
        _has_drawings = False
        with fitz.open(stream=_single_page_bytes, filetype="pdf") as _chk:
            _drawings = _chk[0].get_drawings()
            _has_drawings = bool(_drawings)
        if not _has_drawings:
            logger.info("Page %d has no vector drawings — skipping Camelot", page)

        if _CAMELOT_AVAILABLE and _has_drawings:
            # ── Pass A: full-page camelot for columns, region words for rows ──
            # Results are cached per (session_id, page) so Ghostscript only
            # runs once per page — subsequent region clicks reuse the tables.
            _cache_key = f"{session_id}:{page}" if session_id else None
            with _camelot_cache_lock:
                _cached = _camelot_page_cache.get(_cache_key) if _cache_key else None

            if _cached is not None:
                _tl_cached, _ts_cached = _cached
                logger.info("Camelot cache HIT for %s (lattice=%d stream=%d)",
                            _cache_key, len(_tl_cached), len(_ts_cached))
            else:
                _tl_cached = _ts_cached = None
            def _columns_from_camelot_table(t) -> list[float] | None:
                """Return inter-column x-midpoints restricted to the user's
                horizontal selection. Each camelot column has an [x1, x2]
                extent; we keep only those whose center falls inside the
                selection box, then compute boundaries between the kept
                columns. This way a horizontally-narrow highlight yields a
                narrow result table instead of a wide table mostly padded
                with empty columns.
                """
                cells = getattr(t, "cells", None)
                if not cells:
                    return None
                first_row = cells[0]
                if not first_row or len(first_row) < 2:
                    return None

                # Per-column x-extents, dedup'd by (x1, x2)
                col_ranges = sorted({(round(c.x1, 1), round(c.x2, 1)) for c in first_row})
                if len(col_ranges) < 2:
                    return None

                # Keep columns that meaningfully overlap the user's
                # horizontal selection. A column counts as "selected" when
                # either its center sits inside the box (with a 2pt tol),
                # OR at least half of the column's width is inside the
                # selection. The second condition lets a narrow single-
                # column highlight match a column whose center is slightly
                # outside the box because the user drew over the digits
                # rather than perfectly centred on the column.
                tol = 2.0
                kept: list[tuple[float, float]] = []
                for x1, x2 in col_ranges:
                    center = (x1 + x2) / 2.0
                    overlap = max(0.0, min(x2, rx2) - max(x1, rx1))
                    col_width = max(1.0, x2 - x1)
                    center_in = (rx1 - tol) <= center <= (rx2 + tol)
                    enough_overlap = (overlap / col_width) >= 0.5
                    if center_in or enough_overlap:
                        kept.append((x1, x2))

                if not kept:
                    return None
                if len(kept) == 1:
                    # Exactly one column inside the selection — return an
                    # empty boundaries list so the caller emits a 1-col DF.
                    return []

                # Boundaries are the midpoints between adjacent kept columns
                return [(kept[i][1] + kept[i + 1][0]) / 2.0 for i in range(len(kept) - 1)]

            def _try_fullpage(tables_iter, label: str) -> None:
                best = None
                best_ratio = 0.0
                for t in tables_iter:
                    if t.df is None or t.df.empty:
                        continue
                    bbox = getattr(t, "_bbox", None)
                    if not bbox:
                        continue
                    r = _bbox_overlap_ratio(sel_bbox_bl, bbox)
                    if r > best_ratio:
                        best, best_ratio = t, r
                if best is None or best_ratio <= 0.05:
                    return

                boundaries = _columns_from_camelot_table(best)
                if boundaries is None:
                    # Couldn't recover column structure from camelot's cells
                    # → fall back to using the full table as-is.
                    candidates.append((f"{label}-fullpage", _split_multiline_df(best.df)))
                    logger.info("Region %s-fullpage (raw): shape=%s overlap=%.2f",
                                label, best.df.shape, best_ratio)
                    return
                # boundaries == [] is legal: it means exactly one column of
                # the camelot table sits inside the user's selection. The
                # hybrid extractor will then produce a 1-column DataFrame
                # with one row per visual line in the selection.

                # Re-bucket every word in the user's selection into those columns
                hybrid_df = _words_to_df_with_boundaries(
                    pdf_bytes, page, rx1, ry1_top, rx2, ry2_bot, boundaries,
                )
                if hybrid_df is not None and not hybrid_df.empty:
                    hybrid_df = _split_multiline_df(hybrid_df)
                    candidates.append((f"{label}-hybrid", hybrid_df))
                    logger.info(
                        "Region %s-hybrid: %d cols from camelot, %d rows from region",
                        label, hybrid_df.shape[1], hybrid_df.shape[0],
                    )
                else:
                    candidates.append((f"{label}-fullpage", _split_multiline_df(best.df)))

            def _camelot_safe(flavor, **kwargs):
                """Run camelot with a hard timeout; return [] on timeout/error."""
                def _run():
                    return list(camelot.read_pdf(tmp_path, pages="1", flavor=flavor, **kwargs))
                with concurrent.futures.ThreadPoolExecutor(max_workers=1) as _ex:
                    _fut = _ex.submit(_run)
                    try:
                        return _fut.result(timeout=20)
                    except concurrent.futures.TimeoutError:
                        logger.warning("Camelot %s timed out after 20s", flavor)
                        return []
                    except Exception as exc:
                        logger.debug("camelot %s failed: %s", flavor, exc)
                        return []

            if _tl_cached is not None:
                _try_fullpage(_tl_cached, "camelot-lattice")
            else:
                _tl_cached = _camelot_safe("lattice")
                _try_fullpage(_tl_cached, "camelot-lattice")

            if _ts_cached is not None:
                _try_fullpage(_ts_cached, "camelot-stream")
            else:
                _ts_cached = _camelot_safe("stream", edge_tol=50, row_tol=5)
                _try_fullpage(_ts_cached, "camelot-stream")

            if _cache_key and _camelot_page_cache.get(_cache_key) is None:
                with _camelot_cache_lock:
                    if len(_camelot_page_cache) >= _CAMELOT_CACHE_MAX:
                        _camelot_page_cache.pop(next(iter(_camelot_page_cache)))
                    _camelot_page_cache[_cache_key] = (_tl_cached, _ts_cached)
                logger.info("Camelot cache SET %s (lattice=%d stream=%d)",
                            _cache_key, len(_tl_cached), len(_ts_cached))

            # ── Pass B: region-constrained camelot (existing safety net) ──────
            # If the full-page pass couldn't find anything that overlaps the
            # selection (e.g. the user drew over a table fragment camelot
            # missed at the page level), fall back to the original
            # box-constrained call so we don't regress.
            if not candidates:
                try:
                    tables = camelot.read_pdf(
                        tmp_path, pages="1", flavor="lattice",
                        table_areas=[camelot_area], copy_text=["v"],
                    )
                    for t in tables:
                        if not t.df.empty:
                            candidates.append(("camelot-lattice", _split_multiline_df(t.df)))
                except Exception as exc:
                    logger.debug("camelot lattice region failed: %s", exc)

                try:
                    tables = camelot.read_pdf(
                        tmp_path, pages="1", flavor="stream",
                        table_areas=[camelot_area], edge_tol=50, row_tol=5,
                    )
                    for t in tables:
                        if not t.df.empty:
                            candidates.append(("camelot-stream", _split_multiline_df(t.df)))
                except Exception as exc:
                    logger.debug("camelot stream region failed: %s", exc)
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass

    # ── Pass: direct region OCR (force_ocr — render to image, skip text layer) ─
    _force_ocr_session = False
    if session_id:
        _sd = store.get(session_id)
        _force_ocr_session = bool(getattr(_sd, "force_ocr", False)) if _sd else False

    if _force_ocr_session:
        _rocr_df = _region_ocr_extract(pdf_bytes, eff_page, rx1, ry1_top, rx2, ry2_bot)
        if _rocr_df is not None and not _rocr_df.empty:
            candidates.append(("region-ocr", _split_multiline_df(_rocr_df)))
            logger.info("Region OCR candidate: %dx%d", _rocr_df.shape[0], _rocr_df.shape[1])

    # ── Pass: Docling TableFormer ─────────────────────────────────────────────
    # Always runs on scanned/non-vector pages; also forced when force_ocr=True.
    if (_force_ocr_session or not _has_drawings) and _DOCLING_AVAILABLE:
        _docling_dfs = _docling_table_extract(pdf_bytes, eff_page, rx1, ry1_top, rx2, ry2_bot)
        for _di, _ddf in enumerate(_docling_dfs):
            if not _ddf.empty:
                candidates.append((f"docling-{_di}", _split_multiline_df(_ddf)))
                logger.info("Docling table %d: %dx%d", _di, _ddf.shape[0], _ddf.shape[1])

    # ── Pass: structured_pages words (fastest — already OCR'd by background) ──
    if session_id:
        _sd = store.get(session_id)
        if _sd and _sd.structured_pages and len(_sd.structured_pages) >= page:
            _sp = _sd.structured_pages[page - 1]
            if _sp and _sp.words:
                sp_words = [
                    (w.bbox[0], w.bbox[1], w.bbox[2], w.bbox[3], w.text)
                    for w in _sp.words
                    if w.text.strip()
                    and rx1 <= (w.bbox[0] + w.bbox[2]) / 2 <= rx2
                    and ry1_top <= (w.bbox[1] + w.bbox[3]) / 2 <= ry2_bot
                ]
                if sp_words:
                    df_sp = _words_to_df(sp_words)
                    if df_sp is not None and not df_sp.empty:
                        candidates.append(("structured-pages", df_sp))
                        logger.info("Region: structured_pages → %d words in region", len(sp_words))

    # When force_ocr is active and the re-OCR'd PDF is not ready yet, the
    # pdf_bytes here are the native (corrupt) bytes. Skip pdfplumber/fitz-words
    # to avoid adding garbage candidates; Docling + structured_pages suffice.
    _ocr_pdf_ready = not _force_ocr_session or bool(
        session_id and getattr(store.get(session_id), "searchable_pdf_bytes", None)
    )

    if _ocr_pdf_ready:
        try:
            with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                pg = pdf.pages[eff_page - 1]
                bbox = (rx1, ry1_top, rx2, ry2_bot)
                cropped = pg.within_bbox(bbox)

                for strategy, snap in [("lines_strict", 4), ("lines", 6)]:
                    try:
                        table = cropped.extract_table(table_settings={
                            "vertical_strategy":   strategy,
                            "horizontal_strategy": strategy,
                            "snap_tolerance":      snap,
                            "join_tolerance":      snap,
                            "edge_min_length":     3,
                            "text_x_tolerance":    5,
                            "text_y_tolerance":    5,
                        })
                        if table and any(any(c for c in row) for row in table):
                            rows = [[str(c or "").strip() for c in row] for row in table]
                            ncols = max(len(r) for r in rows) if rows else 0
                            rows = [r + [""] * (ncols - len(r)) for r in rows]
                            candidates.append((f"pdfplumber-{strategy}", pd.DataFrame(rows)))
                    except Exception as exc:
                        logger.debug("pdfplumber %s failed: %s", strategy, exc)
        except Exception as exc:
            logger.debug("pdfplumber open failed: %s", exc)

        try:
            with fitz.open(stream=pdf_bytes, filetype="pdf") as fitz_doc:
                pg = fitz_doc[eff_page - 1]
                clip = fitz.Rect(rx1, ry1_top, rx2, ry2_bot)
                words = pg.get_text("words", clip=clip)
            if words:
                fitz_words = [(w[0], w[1], w[2], w[3], w[4]) for w in words]
                df_w = _words_to_df(fitz_words)
                if df_w is not None and not df_w.empty:
                    candidates.append(("fitz-words", df_w))
        except Exception as exc:
            logger.debug("fitz word-bucket failed: %s", exc)

    if not candidates:
        # Geometry produced something marginal (e.g. 1×N) — fall back to it.
        if geom and geom["cells"]:
            df = _cells_to_flat_df(geom["cells"], geom["n_rows"], geom["n_cols"])
            return {
                "df": df, "cells": geom["cells"],
                "n_rows": geom["n_rows"], "n_cols": geom["n_cols"],
                "source": "geometry-fallback",
            }
        return None

    scored = [(name, df, _score_df(df)) for name, df in candidates]
    scored.sort(key=lambda x: -x[2])
    logger.info(
        "Region candidates: %s → picked %s (%dx%d, score=%.2f)",
        [(n, df.shape, round(s, 2)) for n, df, s in scored],
        scored[0][0], scored[0][1].shape[0], scored[0][1].shape[1], scored[0][2],
    )
    if scored[0][2] <= 0:
        return None
    best_df = scored[0][1]
    return {
        "df": best_df,
        "cells": None,
        "n_rows": int(best_df.shape[0]),
        "n_cols": int(best_df.shape[1]),
        "source": scored[0][0],
    }


# ---------------------------------------------------------------------------
# Region table endpoints
# ---------------------------------------------------------------------------

@router.post("/api/utility/session/{session_id}/extract-table-region")
def extract_table_region(session_id: str, req: RegionRequest):
    if not _EXCEL_AVAILABLE:
        return JSONResponse(status_code=503, content={"error": "pandas not installed"})

    pdf_doc, pdf_bytes = _get_pdf_bytes_for_session(session_id, target_page=req.page)
    if pdf_doc is None:
        return JSONResponse(status_code=404, content={"error": "Session not found. Re-upload."})
    if pdf_bytes is None:
        return JSONResponse(status_code=410, content={"error": "Session files no longer available"})

    try:
        result = _extract_region_table(pdf_bytes, req.page, req.x, req.y, req.width, req.height, session_id)
    except Exception as exc:
        return JSONResponse(status_code=500, content={"error": f"Extraction failed: {exc}"})

    if result is None or result["df"] is None or result["df"].empty:
        return JSONResponse(status_code=404, content={"error": "No text found in selected region"})

    df = result["df"]
    cells = result["cells"]
    n_rows = result["n_rows"]
    n_cols = result["n_cols"]

    rows = [[str(c) for c in row] for row in df.values.tolist()]
    payload = {"rows": rows, "ncols": int(df.shape[1]), "source": result["source"]}
    if cells is not None:
        payload["cells"] = cells
        payload["n_rows"] = n_rows
        payload["n_cols"] = n_cols
    return payload


@router.post("/api/utility/session/{session_id}/extract-table-region/excel")
def extract_table_region_excel(session_id: str, req: RegionRequest):
    if not _EXCEL_AVAILABLE:
        return JSONResponse(status_code=503, content={"error": "pandas not installed"})

    # Use pre-extracted rows when the caller sends them (avoids re-running OCR/Camelot
    # and ensures the export matches exactly what the user sees in the panel, including
    # any sanitization edits applied on the frontend).
    if req.rows is not None and len(req.rows) > 0:
        df = pd.DataFrame(req.rows)
        pdf_doc = store.get(session_id)
        if pdf_doc is None:
            return JSONResponse(status_code=404, content={"error": "Session not found. Re-upload."})
    else:
        pdf_doc, pdf_bytes = _get_pdf_bytes_for_session(session_id, target_page=req.page)
        if pdf_doc is None:
            return JSONResponse(status_code=404, content={"error": "Session not found. Re-upload."})
        if pdf_bytes is None:
            return JSONResponse(status_code=410, content={"error": "Session files no longer available"})

        try:
            result = _extract_region_table(pdf_bytes, req.page, req.x, req.y, req.width, req.height, session_id)
        except Exception as exc:
            return JSONResponse(status_code=500, content={"error": f"Extraction failed: {exc}"})

        if result is None or result["df"] is None or result["df"].empty:
            return JSONResponse(status_code=404, content={"error": "No text found in selected region"})

        df = result["df"]

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Table"

    thin     = Side(style="thin")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="D9E1F2")
    alt_fill = PatternFill("solid", fgColor="F5F7FC")
    hdr_font = Font(bold=True)
    wrap     = Alignment(wrap_text=True, vertical="top")

    col_widths: dict[int, int] = {}

    first_row_is_header = len(df) > 1 and any(
        str(v).strip() and not str(v).replace(".", "").replace("-", "").replace(",", "").strip().isdigit()
        for v in df.iloc[0]
    )
    for ri, row in enumerate(df.itertuples(index=False), start=1):
        is_hdr = first_row_is_header and ri == 1
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=str(val) if val is not None else "")
            cell.border = border
            cell.alignment = wrap
            if is_hdr:
                cell.font = hdr_font
                cell.fill = hdr_fill
            elif ri % 2 == 0:
                cell.fill = alt_fill
            col_widths[ci] = max(col_widths.get(ci, 0), len(str(val or "")))

    for ci, w in col_widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = min(w + 3, 60)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    stem = (pdf_doc.filename or session_id).rsplit(".", 1)[0]
    return Response(
        content=out.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{stem}_region_p{req.page}.xlsx"'},
    )


# ---------------------------------------------------------------------------
# Full-document Excel endpoint
# ---------------------------------------------------------------------------

@router.get("/api/utility/session/{session_id}/excel")
def download_excel(session_id: str):
    if not _EXCEL_AVAILABLE:
        return JSONResponse(
            status_code=503,
            content={"error": "pandas / openpyxl not installed on server."},
        )

    pdf_doc = store.get(session_id)
    if not pdf_doc:
        return JSONResponse(status_code=404, content={"error": "Session not found. Re-upload."})

    needs_ocr = any(p.get("is_ocr") or p.get("is_vector") for p in pdf_doc.page_info)
    _engine = _get_ocr_engine()
    if needs_ocr and pdf_doc.searchable_pdf_bytes is None and _engine is not None:
        logger.info("Excel: generating searchable PDF on-the-fly (image_only) for %s", session_id[-6:])
        try:
            # image_only=True matches the download endpoint and the region
            # extractor — same cache, same behaviour for scanned and
            # vector/UTF-broken PDFs.
            pdf_doc.searchable_pdf_bytes = make_searchable_pdf(
                pdf_doc, _engine, image_only=True,
            )
            pdf_doc._searchable_cache_key = (False, True)  # noqa: SLF001
        except Exception as exc:
            logger.warning("Excel: searchable PDF generation failed: %s", exc)

    temp_path: str | None = None
    try:
        if pdf_doc.searchable_pdf_bytes:
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as fh:
                fh.write(pdf_doc.searchable_pdf_bytes)
                temp_path = fh.name
            pdf_path = temp_path
        elif pdf_doc.pdf_path:
            pdf_path = str(pdf_doc.pdf_path)
        elif pdf_doc.pdf_bytes:
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as fh:
                fh.write(pdf_doc.pdf_bytes)
                temp_path = fh.name
            pdf_path = temp_path
        else:
            return JSONResponse(status_code=410, content={"error": "Session files no longer available"})

        try:
            with fitz.open(pdf_path) as _probe:
                total_pages = len(_probe)
        except Exception as exc:
            return JSONResponse(status_code=500, content={"error": f"Cannot open PDF: {exc}"})

        combined: list = []
        if _CAMELOT_AVAILABLE:
            tables_lattice: list = []
            try:
                tl = camelot.read_pdf(pdf_path, flavor="lattice", pages="all")
                tables_lattice = list(tl)
                logger.info("Lattice: %d table(s)", len(tables_lattice))
            except Exception as exc:
                logger.warning("Lattice failed: %s", exc)

            lattice_pages: set[int] = {t.page for t in tables_lattice}

            tables_stream: list = []
            uncovered = [p for p in range(1, total_pages + 1) if p not in lattice_pages]
            if uncovered:
                try:
                    ts = camelot.read_pdf(
                        pdf_path, flavor="stream",
                        pages=",".join(str(p) for p in uncovered),
                        edge_tol=50, row_tol=5,
                    )
                    tables_stream = list(ts)
                    logger.info("Stream on %d page(s): %d table(s)", len(uncovered), len(tables_stream))
                except Exception as exc:
                    logger.warning("Stream failed: %s", exc)

            combined = list(tables_lattice)
            for st in tables_stream:
                if not any(
                    lt.page == st.page and _bbox_overlap_ratio(lt._bbox, st._bbox) > 0.5
                    for lt in tables_lattice
                ):
                    combined.append(st)
            combined.sort(key=lambda t: t.page)
        else:
            logger.info("camelot not available — using PyMuPDF fallback only")

        camelot_pages: set[int] = {t.page for t in combined}
        logger.info("Combined tables: %d across %d page(s)", len(combined), len(camelot_pages))

        page_text_dfs: dict[int, "pd.DataFrame"] = {}
        try:
            with fitz.open(pdf_path) as fitz_doc:
                for page_idx in range(total_pages):
                    page_num = page_idx + 1
                    if page_num in camelot_pages:
                        continue
                    try:
                        words = fitz_doc[page_idx].get_text("words")
                        df = _words_to_df(words)
                        if df is not None and not df.empty:
                            page_text_dfs[page_num] = df
                    except Exception as exc:
                        logger.warning("Fallback page %d: %s", page_num, exc)
        except Exception as exc:
            logger.warning("Fallback open failed: %s", exc)

        if not combined and not page_text_dfs:
            return JSONResponse(status_code=404, content={"error": "No content found in this PDF."})

        all_page_nums = sorted(camelot_pages | set(page_text_dfs.keys()))
        tables_by_page: dict[int, list] = {}
        for t in combined:
            tables_by_page.setdefault(t.page, []).append(t)

        GAP_ROWS     = 3
        HDR_ROWS     = 2
        display_name = pdf_doc.filename or f"{session_id}.pdf"

        sheet_names_used: set[str] = set()
        sheet_regions: dict[str, list[tuple[int, int, int]]] = {}

        def _unique_sheet(name: str) -> str:
            base = name[:31]
            if base not in sheet_names_used:
                sheet_names_used.add(base)
                return base
            for i in range(2, 9999):
                cand = f"{base[:28]}_{i}"
                if cand not in sheet_names_used:
                    sheet_names_used.add(cand)
                    return cand
            return base

        excel_buf = io.BytesIO()
        wrote_any = False

        with pd.ExcelWriter(excel_buf, engine="openpyxl") as writer:
            for page_num in all_page_nums:
                page_tables = tables_by_page.get(page_num, [])
                page_tables = [t for t in page_tables if t.df is not None and not t.df.empty]
                page_df     = page_text_dfs.get(page_num)

                if not page_tables and (page_df is None or page_df.empty):
                    continue

                sheet = _unique_sheet(f"Page {page_num}")
                regions: list[tuple[int, int, int]] = []

                pd.DataFrame([[display_name]]).to_excel(
                    writer, sheet_name=sheet, index=False, header=False, startrow=0,
                )

                start_row = HDR_ROWS

                if page_tables:
                    for table in page_tables:
                        df = _split_multiline_df(table.df)
                        df.to_excel(
                            writer, sheet_name=sheet,
                            index=False, header=False, startrow=start_row,
                        )
                        regions.append((start_row + 1, len(df), len(df.columns)))
                        start_row += len(df) + GAP_ROWS
                elif page_df is not None and not page_df.empty:
                    page_df.to_excel(
                        writer, sheet_name=sheet,
                        index=False, header=False, startrow=start_row,
                    )
                    regions.append((start_row + 1, len(page_df), len(page_df.columns)))

                sheet_regions[sheet] = regions
                wrote_any = True

            if not wrote_any:
                s = _unique_sheet("Empty")
                pd.DataFrame([[display_name], ["No extractable content"]]).to_excel(
                    writer, sheet_name=s, index=False, header=False,
                )

        formatted = _apply_excel_formatting(excel_buf.getvalue(), sheet_regions)

        stem = (pdf_doc.filename or session_id).rsplit(".", 1)[0]
        return Response(
            content=formatted,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{stem}_tables.xlsx"'},
        )

    finally:
        if temp_path:
            try:
                os.unlink(temp_path)
            except Exception:
                pass
