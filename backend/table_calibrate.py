"""Interactive table calibration.

Flow used by the frontend /table-scrape page:

  1. Upload a PDF or image via the normal /api/utility/process endpoint.
  2. GET  /api/table/detect/{session_id}/{page}
        → returns the page's pixel size plus the auto-detected column and
          row separator lines (in the SAME pixel space as the /page PNG).
  3. The user drags / adds / deletes those lines.
  4. POST /api/table/scrape/{session_id}/{page}   { columns:[x], rows:[y] }
        → returns the table (list of rows) bucketed by the user's lines.
  5. POST /api/table/excel/{session_id}/{page}    { columns:[x], rows:[y] }
        → returns the same table as an .xlsx download.

Coordinates everywhere are pixels of the page rendered at RENDER_DPI (150),
which is exactly what GET /api/utility/page/{session_id}/{page} serves, so
the frontend overlay lines up 1:1 with the displayed image.
"""

import io

import fitz
import numpy as np
from fastapi import APIRouter, Body
from fastapi.responses import JSONResponse, Response

from pdf_store import store
from ocr_lazy import get_ocr_engine

router = APIRouter()

RENDER_DPI = 150
_SCALE = RENDER_DPI / 72.0


# ---------------------------------------------------------------------------
# Word extraction (native text layer if present, else OCR) in pixel space
# ---------------------------------------------------------------------------

def _page_words_px(pdf_doc, page_idx: int):
    """Return (words, width_px, height_px).

    words: list of {text, x0, y0, x1, y1} in pixels of the RENDER_DPI image.
    Prefers the native text layer; falls back to OCR of the rendered page.
    """
    page = pdf_doc.doc[page_idx]
    rect = page.rect
    width_px = int(round(rect.width * _SCALE))
    height_px = int(round(rect.height * _SCALE))

    words = []
    try:
        native = page.get_text("words") or []
    except Exception:
        native = []
    # get_text("words") → (x0, y0, x1, y1, "word", block, line, word_no)
    for w in native:
        text = (w[4] or "").strip()
        if not text:
            continue
        words.append({
            "text": text,
            "x0": w[0] * _SCALE, "y0": w[1] * _SCALE,
            "x1": w[2] * _SCALE, "y1": w[3] * _SCALE,
        })

    if words:
        return words, width_px, height_px

    # No native text → OCR the rendered page (2x upscale helps small text).
    engine = get_ocr_engine()
    if engine is None:
        return [], width_px, height_px
    mat = fitz.Matrix(_SCALE * 2, _SCALE * 2)  # render at 2x for OCR accuracy
    pix = page.get_pixmap(matrix=mat)
    try:
        img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
        if pix.n == 4:
            img = img[:, :, :3]
        img = np.ascontiguousarray(img)
        res = engine.ocr(img, cls=True)
    finally:
        pix = None
    if res and res[0]:
        for box, (text, _score) in res[0]:
            t = (text or "").strip()
            if not t:
                continue
            xs = [p[0] for p in box]
            ys = [p[1] for p in box]
            # divide by 2 to map the 2x-OCR coords back to RENDER_DPI space
            words.append({
                "text": t,
                "x0": min(xs) / 2, "y0": min(ys) / 2,
                "x1": max(xs) / 2, "y1": max(ys) / 2,
            })
    return words, width_px, height_px


# ---------------------------------------------------------------------------
# Auto-detection of column / row separator lines
# ---------------------------------------------------------------------------

def _detect_columns(words, width_px):
    """Find vertical separator x-positions via whitespace-gap analysis.

    Marks every x column covered by any word, then places a separator in the
    middle of each sufficiently wide uncovered gap.
    """
    if not words or width_px <= 0:
        return []
    cover = np.zeros(width_px + 1, dtype=bool)
    heights = []
    for w in words:
        x0 = max(0, int(w["x0"]))
        x1 = min(width_px, int(w["x1"]))
        if x1 > x0:
            cover[x0:x1] = True
        heights.append(w["y1"] - w["y0"])
    med_h = float(np.median(heights)) if heights else 12.0
    # A real column gap is wider than an inter-word space. Use ~1.2x the text
    # height as the whitespace threshold (tunable, and the user calibrates).
    gap_min = max(med_h * 1.2, 14)

    seps = []
    x = 0
    # skip the leading margin
    while x < width_px and not cover[x]:
        x += 1
    while x < width_px:
        if not cover[x]:
            start = x
            while x < width_px and not cover[x]:
                x += 1
            end = x
            # ignore the trailing margin (gap that runs to the edge)
            if end < width_px and (end - start) >= gap_min:
                seps.append(int((start + end) / 2))
        else:
            x += 1
    return seps


def _detect_rows(words, height_px):
    """Cluster words into text lines; separator = midpoint between lines."""
    if not words:
        return []
    heights = [w["y1"] - w["y0"] for w in words]
    med_h = float(np.median(heights)) if heights else 12.0
    tol = med_h * 0.6
    centers = sorted(((w["y0"] + w["y1"]) / 2) for w in words)
    lines = []
    cur = [centers[0]]
    for c in centers[1:]:
        if c - cur[-1] <= tol:
            cur.append(c)
        else:
            lines.append(sum(cur) / len(cur))
            cur = [c]
    lines.append(sum(cur) / len(cur))
    seps = [int((lines[i] + lines[i + 1]) / 2) for i in range(len(lines) - 1)]
    return seps


# ---------------------------------------------------------------------------
# Scrape given calibrated lines
# ---------------------------------------------------------------------------

def _scrape_with_lines(words, width_px, height_px, columns, rows):
    """Bucket words into a 2-D table using the given separator lines."""
    col_edges = [0] + sorted(int(c) for c in columns) + [width_px]
    row_edges = [0] + sorted(int(r) for r in rows) + [height_px]
    n_cols = len(col_edges) - 1
    n_rows = len(row_edges) - 1
    grid = [[[] for _ in range(n_cols)] for _ in range(n_rows)]

    def _bucket(edges, v):
        for i in range(len(edges) - 1):
            if edges[i] <= v < edges[i + 1]:
                return i
        return len(edges) - 2  # clamp into last cell

    for w in words:
        cx = (w["x0"] + w["x1"]) / 2
        cy = (w["y0"] + w["y1"]) / 2
        ci = _bucket(col_edges, cx)
        ri = _bucket(row_edges, cy)
        grid[ri][ci].append((w["x0"], w["text"]))

    table = []
    for r in range(n_rows):
        row_out = []
        for c in range(n_cols):
            parts = [t for _, t in sorted(grid[r][c], key=lambda p: p[0])]
            row_out.append(" ".join(parts).strip())
        if any(cell for cell in row_out):
            table.append(row_out)
    return table


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@router.get("/api/table/detect/{session_id}/{page}")
def detect_lines(session_id: str, page: int):
    pdf_doc = store.get(session_id)
    if not pdf_doc:
        return JSONResponse(status_code=404, content={"error": "Session not found. Re-upload."})
    if page < 1 or page > pdf_doc.page_count:
        return JSONResponse(status_code=400, content={"error": f"Invalid page. PDF has {pdf_doc.page_count} pages."})

    words, w_px, h_px = _page_words_px(pdf_doc, page - 1)
    columns = _detect_columns(words, w_px)
    rows = _detect_rows(words, h_px)
    return {
        "width": w_px,
        "height": h_px,
        "columns": columns,
        "rows": rows,
        "word_count": len(words),
    }


@router.post("/api/table/scrape/{session_id}/{page}")
def scrape_lines(session_id: str, page: int, body: dict = Body(...)):
    pdf_doc = store.get(session_id)
    if not pdf_doc:
        return JSONResponse(status_code=404, content={"error": "Session not found. Re-upload."})
    if page < 1 or page > pdf_doc.page_count:
        return JSONResponse(status_code=400, content={"error": f"Invalid page. PDF has {pdf_doc.page_count} pages."})

    words, w_px, h_px = _page_words_px(pdf_doc, page - 1)
    columns = body.get("columns", []) or []
    rows = body.get("rows", []) or []
    table = _scrape_with_lines(words, w_px, h_px, columns, rows)
    return {"rows": table, "n_cols": (len(columns) + 1), "n_rows": len(table)}


@router.post("/api/table/excel/{session_id}/{page}")
def scrape_excel(session_id: str, page: int, body: dict = Body(...)):
    pdf_doc = store.get(session_id)
    if not pdf_doc:
        return JSONResponse(status_code=404, content={"error": "Session not found. Re-upload."})
    if page < 1 or page > pdf_doc.page_count:
        return JSONResponse(status_code=400, content={"error": f"Invalid page. PDF has {pdf_doc.page_count} pages."})

    import openpyxl

    words, w_px, h_px = _page_words_px(pdf_doc, page - 1)
    columns = body.get("columns", []) or []
    rows = body.get("rows", []) or []
    first_row_header = bool(body.get("first_row_header", False))
    table = _scrape_with_lines(words, w_px, h_px, columns, rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Table"
    if first_row_header and table:
        from openpyxl.styles import Font
        ws.append(table[0])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for r in table[1:]:
            ws.append(r)
    else:
        for r in table:
            ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)

    base = (pdf_doc.filename or session_id).rsplit(".", 1)[0]
    fname = f"{base}_table.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------------------------------------------------------------------------
# All-pages: apply ONE set of calibrated lines to every page of the PDF
# ---------------------------------------------------------------------------

def _scaled_lines(columns, rows, ref_w, ref_h, page_w, page_h):
    """Scale the calibrated lines (defined in the reference page's pixel space)
    into a target page's pixel space. For a uniform-size PDF this is a no-op;
    for pages of differing size it keeps the grid proportional."""
    sx = (page_w / ref_w) if ref_w else 1.0
    sy = (page_h / ref_h) if ref_h else 1.0
    return [c * sx for c in columns], [r * sy for r in rows]


def _scrape_all_pages(pdf_doc, columns, rows, ref_w, ref_h):
    """Scrape every page with the same (scaled) calibrated lines.
    Returns list of (page_number, table_rows)."""
    out = []
    for i in range(pdf_doc.page_count):
        words, w_px, h_px = _page_words_px(pdf_doc, i)
        cols_i, rows_i = _scaled_lines(columns, rows, ref_w or w_px, ref_h or h_px, w_px, h_px)
        table = _scrape_with_lines(words, w_px, h_px, cols_i, rows_i)
        out.append((i + 1, table))
    return out


@router.post("/api/table/scrape-all/{session_id}")
def scrape_all(session_id: str, body: dict = Body(...)):
    """Preview: scrape every page with the calibrated lines. Returns rows with
    a leading page number so the frontend can show the whole-PDF result."""
    pdf_doc = store.get(session_id)
    if not pdf_doc:
        return JSONResponse(status_code=404, content={"error": "Session not found. Re-upload."})
    columns = body.get("columns", []) or []
    rows = body.get("rows", []) or []
    ref_w = float(body.get("ref_width") or 0)
    ref_h = float(body.get("ref_height") or 0)

    pages = _scrape_all_pages(pdf_doc, columns, rows, ref_w, ref_h)
    combined = []
    for page_no, table in pages:
        for r in table:
            combined.append([str(page_no)] + r)
    return {
        "rows": combined,
        "n_cols": len(columns) + 2,     # page col + data cols
        "n_rows": len(combined),
        "page_count": pdf_doc.page_count,
    }


@router.post("/api/table/excel-all/{session_id}")
def scrape_excel_all(session_id: str, body: dict = Body(...)):
    """Download the WHOLE PDF as one Excel: the calibrated lines are applied to
    every page and all rows are stacked with a leading Page column."""
    pdf_doc = store.get(session_id)
    if not pdf_doc:
        return JSONResponse(status_code=404, content={"error": "Session not found. Re-upload."})

    import openpyxl
    from openpyxl.styles import Font

    columns = body.get("columns", []) or []
    rows = body.get("rows", []) or []
    ref_w = float(body.get("ref_width") or 0)
    ref_h = float(body.get("ref_height") or 0)
    first_row_header = bool(body.get("first_row_header", False))

    pages = _scrape_all_pages(pdf_doc, columns, rows, ref_w, ref_h)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Table"

    header_written = False
    total = 0
    for page_no, table in pages:
        if not table:
            continue
        start = 0
        if first_row_header:
            if not header_written:
                ws.append(["Page"] + table[0])
                for cell in ws[ws.max_row]:
                    cell.font = Font(bold=True)
                header_written = True
            # On every page the first row is the repeated column header — skip it.
            start = 1
        for r in table[start:]:
            ws.append([str(page_no)] + r)
            total += 1

    if total == 0 and not header_written:
        ws.append(["Page", "(no rows found on any page)"])

    buf = io.BytesIO()
    wb.save(buf)

    base = (pdf_doc.filename or session_id).rsplit(".", 1)[0]
    fname = f"{base}_all_pages.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{fname}"',
            "X-Table-Rows": str(total),
            "X-Table-Pages": str(pdf_doc.page_count),
            "Access-Control-Expose-Headers": "X-Table-Rows,X-Table-Pages",
        },
    )
